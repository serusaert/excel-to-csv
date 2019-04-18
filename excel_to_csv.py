# \\V-FS4\nc\sas\macros\utilities\python\csv-remove-newline\excel_to_csv.py
# Kevin Beverly 2018-06-24

# Save XLS worksheet as CSV with newlines removed as UTF-8 (non-translatable chars print as "?")
# Argument 1: Full path to Excel file
# Argument 2: Name of worksheet to export
# Argument 3: Full path to CSV output file

# Call from SAS:
#  %let spreadsheet = \\v-fs4\nc\sas\p-15\P15_2013\inputs\conmed\p15_conmed_coding_22JUN2018_kb.xlsx;
#  %let worksheet = Not coded; ** has newlines embedded in cells;
#  %let pyprog = \\V-FS4\nc\sas\macros\utilities\python\csv-remove-newline\excel_to_csv.py;
#  %let outfile = \\v-fs4\nc\sas\p-15\P15_2013\inputs\conmed\test1.csv;
#  x C:\Python\Python36\python.exe "&pyprog" "&spreadsheet" "&worksheet" "&outfile" ;
#  proc import datafile="&outfile" out=new_coded dbms=csv replace;
#        GUESSINGROWS=max;
#        GETNAMES=YES;
#  run;

import openpyxl
import csv, sys, os

# Get command line arguments
#print ("args: ", len(sys.argv), sys.argv[1], sys.argv[2], sys.argv[3] )
if len(sys.argv) != 4:
    print ( "\n", 'Incorrect number of input parameters found!', "\n", 'Usage: ', sys.argv[0], ' [full path to Excel file, worksheet name, full path to output CSV file]' )
    sys.exit() 

if len(sys.argv) == 4 and os.path.isfile(sys.argv[1]):
    excel_path = str(sys.argv[1])
    excel_path = excel_path.strip()
else:
    print ( "\n", 'Excel file not found!', "\n", 'Usage: ', sys.argv[0], ' [full path to Excel file, worksheet name, full path to output CSV file]', "\n", 'Param: ', sys.argv[1] )
    sys.exit() 

if len(sys.argv) == 4 and str(sys.argv[2]):
    worksheet = str(sys.argv[2])
    worksheet = worksheet.strip()
else:
    print ( "\n", 'Excel worksheet name not found!', "\n", 'Usage: ', sys.argv[0], ' [full path to Excel file, worksheet name, full path to output CSV file]', "\n", 'Param: ', sys.argv[2] )
    sys.exit() 

if len(sys.argv) == 4: # and os.path.isfile(sys.argv[3]):
    csv_path = str(sys.argv[3])
    csv_path = csv_path.strip()

# Export CSV file from Excel worksheet
wb = openpyxl.load_workbook(excel_path)
sh = wb.get_sheet_by_name(worksheet)
row_count = sh.max_row
column_count = sh.max_column
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
   c = csv.writer(f)
   for r in sh.rows:
      c.writerow([cell.value for cell in r])

# Replace newlines with spaces within cells
new_rows = [] # a holder for our modified rows when we make them
with open(csv_path, 'rt', newline='') as f:
    reader = csv.reader(f) # pass the file to our csv reader
    for row in reader:     # iterate over the rows in the file
        new_row = row      # at first, just copy the row
        new_row = [ x.replace('\n', ' ') for x in new_row ] # make the substitutions
        new_rows.append(new_row) # add the modified rows

# Overwrite the CSV file with the modified rows
with open(csv_path, 'wt', newline='') as f:
    writer = csv.writer(f)
    writer.writerows(new_rows)

print ( "Finished: row_count=", row_count, "; column_count=", column_count )
